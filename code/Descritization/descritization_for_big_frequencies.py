#!/usr/bin/env python
#descritization_AUIDis.py

""" Performs AUI-Dis based equal frequency binning.

    AUI-Dis uses iterative binning to find potential number of bins and automatically 
    adapts the bin boundaries in order to ensure that all occurrences of the same 
    value are binned together, while maximizing the bin coherence.

    Attributes:
        input_file: Location where input required resources are stored.
        output_directory: Location where descritization results and plots are stored.
        sheet_name: Sheet Name where the dist123 or maxDist or theta is present.
        
"""

from __future__ import division
from copy import deepcopy
import argparse, csv, glob, math, os, operator, re, time

from itertools import groupby
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import itertools, ntpath, xlrd, xlsxwriter 
import pandas as pd, numpy as np, matplotlib.pyplot as plt
from pandas import ExcelWriter
from pandas import ExcelFile
  

__author__ = "Venkata Sarika Kondra"

__version__ = "1.0.1"
__maintainer__ = "Venkata Sarika Kondra"
__email__ = "c00219805@louisiana.edu"

parser = argparse.ArgumentParser(description='Adaptive Equal Frequency Binning.')

parser.add_argument('--sheet_name', '-s', metavar='sheet_name', default='dist123-sample4',
                   help='Sheet Name where the dist123 or maxDist or theta is present.')
parser.add_argument('--input_file', dest = 'input_file', metavar='input_file', 
                    default="//home//linc//c00219805//Research//Descritization/input/new_samples_sarika_theta-dist-distribution260118.xlsx" ,
                    #default="//home//linc//c00219805//Research///Descritization/input/new_samples_sarika_theta-dist-distribution290118_round_off_2.xlsx" ,
                    #default="//home//linc//c00219805//Research///Sumi_descritization/Sumi_theta-dist-distribution.xlsx" ,
                    help='Location where input required resources are stored')
parser.add_argument('--output_directory', dest = 'output_directory', metavar='output_directory', 
                    #default="//home//linc//c00219805//Research//Descritization/" ,
                    default="//home//linc//c00219805//Research//Sumi_descritization/" ,
                    help='Location where descritization results and plots are stored')


def read_df_from_csv(file_path,sep,header): 
    """Read a dataframe from the CSV file."""
    return pd.DataFrame.from_csv(file_path,sep = sep,header = header)
def discretize(data,type,diffFactor):
    '''Returns number of discretization levels and bin boundaries''' 
    sizeOfSample = data['freq'].sum()
    x = data[type].values
    frequency = data['freq'].values
    maxNumOfBins = int(round(math.log(sizeOfSample,2) + 1,0))
    print 'Max allowed bins: ', maxNumOfBins
    recalculateBins = []
    
    binBoundariesDict = {}
    binFrequenciesDict = {}
    binExpectedFreqDict = {}
    binTempVariance = {}
    firstIter = True
    epoch = 0
    while(firstIter or recalculateBins):
        epoch =epoch+1
        print '-------------------------------------------------------------------------------------------------------------'
        print 'epoch ' , epoch, recalculateBins
        iterRange = range(2,maxNumOfBins + 1)
        #iterRange.remove(27)
        #iterRange.remove(26)
        if not firstIter:
            iterRange = recalculateBins
        for bin in iterRange:
                if not firstIter:
                    recalculateBins.remove(bin)
                expFreq = sizeOfSample/bin
                binExpectedFreqDict[bin] = expFreq
                print 'bin:' , bin, 'expFreq:', expFreq
                if firstIter:
                    binFrequencies,binBoundaries = np.histogram(x,bins= bin,weights = frequency)
        
                else:
                    binFrequencies,binBoundaries = np.histogram(x,bins= binBoundariesDict[bin],weights = frequency)
                #print 'started at..',binFrequencies,binBoundaries
                for i in range(1,bin):
                    #while the bin frequency is less than expected frequency include next all x's until bin frequency becomes greater than expected frequency
                    if (binFrequencies[i-1] < expFreq):
                        while((binFrequencies[i-1] < expFreq)):
                                    tempBinBoundaries,tempBinFrequencies = deepcopy(binBoundaries),deepcopy(binFrequencies)
                                    tempBinBoundaries[i] = tempBinBoundaries[i]+float(diffFactor)
                                    #current bin higher limit cannot exceed next bin's lower limit
                                    if(tempBinBoundaries[i] < tempBinBoundaries[i+1]):
                                        tempBinFrequencies,tempBinBoundaries = np.histogram(x,bins= tempBinBoundaries,weights = frequency)
                                        if(tempBinFrequencies[i-1] < expFreq):
                                            binBoundaries = tempBinBoundaries
                                            binBoundariesDict[bin]=binBoundaries
                                            binFrequencies = tempBinFrequencies
                                            binFrequenciesDict[bin] = binFrequencies
                                            binTempVariance['bin'+str(bin)+str(i)] = abs(tempBinFrequencies[i-1] - expFreq)
                                        else:
                                            devMax = abs(tempBinFrequencies[i-1] - expFreq)
                                            if 'bin'+str(bin)+str(i) not in binTempVariance:
                                                binTempVariance['bin'+str(bin)+str(i)] = devMax
                                            if (devMax <= binTempVariance['bin'+str(bin)+str(i)]):
                                                binBoundaries = tempBinBoundaries
                                                binBoundariesDict[bin]= binBoundaries
                                                binFrequencies = tempBinFrequencies
                                                binFrequenciesDict[bin] = binFrequencies
                                            break
                                    else:
                                        if bin not in recalculateBins:
                                            recalculateBins.append(bin)
                                        break
                    else:
                        while((binFrequencies[i-1] > expFreq)):
                                    tempBinBoundaries,tempBinFrequencies = deepcopy(binBoundaries),deepcopy(binFrequencies)
                                    tempBinBoundaries[i] = tempBinBoundaries[i]-float(diffFactor)
                                    #current bin higher limit cannot exceed its own lower limit
                                    if(tempBinBoundaries[i-1] < tempBinBoundaries[i]):
                                        tempBinFrequencies,tempBinBoundaries = np.histogram(x,bins= tempBinBoundaries,weights = frequency)
                                        if(tempBinFrequencies[i-1] > expFreq):
                                            binBoundaries = tempBinBoundaries
                                            binBoundariesDict[bin]=binBoundaries
                                            binFrequencies = tempBinFrequencies
                                            binFrequenciesDict[bin] = binFrequencies
                                            binTempVariance['bin'+str(bin)+str(i)] = abs(tempBinFrequencies[i-1] - expFreq)
                                        else:
                                            devMinus = abs(tempBinFrequencies[i-1] - expFreq)
                                            if 'bin'+str(bin)+str(i) not in binTempVariance:
                                                binTempVariance['bin'+str(bin)+str(i)] = devMinus
                                            if (devMinus <= binTempVariance['bin'+str(bin)+str(i)]):
                                                binBoundaries = tempBinBoundaries
                                                binBoundariesDict[bin]= binBoundaries
                                                binFrequencies = tempBinFrequencies
                                                binFrequenciesDict[bin] = binFrequencies
                                            break
                                    else:
                                        if bin not in recalculateBins:
                                            recalculateBins.append(bin)
                                        break
                end_sum = 0
                for f in binFrequencies:
                        end_sum += (f/sizeOfSample)*(f- binExpectedFreqDict[bin])**2
                
                print 'Bin Frequencies: ',binFrequencies
                #print "Bin Boundaries: ",binBoundaries
        firstIter = False
    
    #Calculating Optimal bin
    binVariances = {}
    for bin, freq in binFrequenciesDict.iteritems():
        sum = 0
        for f in freq:
            sum += (f/sizeOfSample)*(f- binExpectedFreqDict[bin])**2
        binVariances[bin] = sum
    optimalBin = min(binVariances, key=binVariances.get)
    print 'Bin Boundaries: ',binBoundariesDict
    print 'optimalBin ',optimalBin
    print 'Optimal Bin Boundaries:',binBoundariesDict[optimalBin], 'Optimal Bin Frequencies: ',binFrequenciesDict[optimalBin], 'Expected Frequency: ',binExpectedFreqDict[optimalBin]
    print sorted(binVariances.items(), key=operator.itemgetter(1))
    #print binBoundariesDict
    return binBoundariesDict[optimalBin], binFrequenciesDict[optimalBin],binVariances,maxNumOfBins


if __name__ == '__main__':
    """Executable code from command line."""
    args = parser.parse_args()
    
    ######Perform Discretization######
    
    type = ''
    xlabel = ''
    title = ''
    step = 0.01
    if "theta" in args.sheet_name:
        #Theta
        #df = read_df_from_csv(input_files+'//allthetas.csv',',',0)
        type = 'theta'
        xlabel = 'Theta in Degrees'
        title = 'Theta Distribution'
    else:
        #Distance    
        type = 'maxDist'
        xlabel = 'Length in Angstroms'
        title = 'Distribution of Length'
        step = 0.025
    df = pd.read_excel(args.input_file, sheetname=args.sheet_name)
    max_freq = int(max(df['freq'].values))
    max_x = int(max(df[type].values))

    #Plotting
    optimal_bins, frequencies,all_variances,maxNumOfBins= discretize(df,type,step)
    text =  [round(b,2) for b in optimal_bins]
    print text
    fig , ax = plt.subplots()
    ax.vlines(optimal_bins,ymin = 0,ymax = max_freq +10000,label = optimal_bins,color = 'y')
    plt.scatter(df[type].values, df['freq'].values, s=50)
    plt.title(title)
    plt.text(max_x/2 + 1,(max_freq +10000)/2,r'$BinBoundaries $' + '\n' +str(text),color='red', fontsize=8)
    plt.xlabel(type)
    plt.ylabel('Frequency')
    plt.savefig(args.output_directory +args.sheet_name +'.png', dpi=200)
    #Saving to excel
    writer = pd.ExcelWriter(args.output_directory+args.sheet_name+'.xlsx', engine='xlsxwriter')
    df.to_excel(writer, sheet_name=args.sheet_name)
    worksheet = writer.sheets[args.sheet_name]
    worksheet.insert_image('E5', args.output_directory +args.sheet_name +'.png')

    plt.show() #Shows the plot on the screen. Comment it if not required.

    #Total Variance plotting
    df_variance = pd.DataFrame(all_variances.items() , columns= ['bin','variance'])    
    plt.xticks(all_variances.keys())
    plt.plot(all_variances.keys(), all_variances.values(), '-o')
    plt.xlabel('Bin#')
    plt.ylabel('Variance')
    plt.title('Bin# - Variance Plot (' + args.sheet_name + ')')
    plt.savefig(args.output_directory +'all_variances_' +args.sheet_name +'.png', dpi=200)
    #plt.show() #Shows the plot on the screen. Comment it if not required.

    if os.path.isfile(args.output_directory+'all_variance.xlsx'):
        book = load_workbook(args.output_directory+'all_variance.xlsx')
        #book = xlsxwriter.Workbook(output_directory+'all_variance.xlsx')
        #workbook.add_worksheet()
        writer = pd.ExcelWriter(args.output_directory+'all_variance.xlsx', engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_variance.to_excel(writer, args.sheet_name)    
        ws = book.worksheets[len(writer.sheets) -1]
        ws.add_image(Image(args.output_directory+ 'all_variances_' +args.sheet_name +'.png'), 'E5')
        book.save(args.output_directory+'all_variance.xlsx')

        
    else:
        writer = pd.ExcelWriter(args.output_directory+'all_variance.xlsx',  engine='xlsxwriter')
        df_variance.to_excel(writer, sheet_name=args.sheet_name)
        worksheet = writer.sheets[args.sheet_name]
        worksheet.insert_image('E5', args.output_directory+ 'all_variances_' +args.sheet_name +'.png')
        #writer.save()

   

